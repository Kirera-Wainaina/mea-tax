import openpyxl
from openpyxl.drawing.image import Image
import os
import shutil
import math

def open_workbook():
    workbook = openpyxl.load_workbook(filename='./static/records.xlsx')
    return workbook

def get_records_worksheet(workbook):
    return workbook['PAYE2023']

def iterate_through_rows(worksheet, 
        min_row, max_row,
        min_col, max_col):
    for row in worksheet.iter_rows(
            min_row=min_row, max_row=max_row, 
            min_col=min_col, max_col=max_col):
        handle_employee_details(row)

def handle_employee_details(details):
    employee_name = details[1].value
    file_path = create_employee_tax_file(employee_name)
    employee_workbook = load_employee_workbook(file_path)
    p9_sheet = employee_workbook.active

    add_employee_name_to_their_worksheet(p9_sheet, employee_name)
    add_employee_pin_to_their_worksheet(p9_sheet, details[0].value)
    add_kra_logo_to_employee_worksheet(p9_sheet)
    add_financial_details_to_employee_worksheet(p9_sheet, details)

    employee_workbook.save(filename=file_path)
    return True

def create_employee_tax_file(name):
    template_path = '{cwd}/static/p9.xlsx'.format(cwd=os.getcwd())
    employee_file_path = '{cwd}/repo/{name}.xlsx'.format(
        cwd=os.getcwd(), name=name)
    shutil.copy(template_path, employee_file_path)
    return employee_file_path

def load_employee_workbook(file_path):
    return openpyxl.load_workbook(file_path)

def add_employee_name_to_their_worksheet(worksheet, name):
    worksheet['D12'] = name
    return True

def add_employee_pin_to_their_worksheet(worksheet, pin):
    worksheet['L14'] = pin
    return

def add_kra_logo_to_employee_worksheet(worksheet):
    image = Image('{cwd}/static/p9_logo.png'.format(cwd=os.getcwd()))
    worksheet.add_image(image, 'H2')
    return

def add_financial_details_to_employee_worksheet(worksheet, details):
    # remove name and pin from details
    # remain with salary and paye from Jan to Dec
    financial_details = details[2:26]
    for index, item in enumerate(financial_details):
        if index % 2 == 0:
            add_salary_to_employee_worksheet(worksheet, index, item.value)
        else:
            add_tax_to_employee_worksheet(worksheet, index, item.value)
    return

def add_salary_to_employee_worksheet(worksheet, index, item):
    salary_column, start_row = 'C', 26
    current_row = math.floor(start_row + index/2)
    if item == '-' or type(item) == type(None):
        item = 0
    worksheet['{column}{row}'.format(column=salary_column, row=current_row)] = item
    return

def add_tax_to_employee_worksheet(worksheet, index, item):
    tax_column, start_row = 'O', 26
    current_row = math.floor(start_row + (index-1)/2)

    if item == '-' or type(item) == type(None):
        item = 0

    worksheet['{column}{row}'.format(column=tax_column, row=current_row)] = item
    adjust_chargeable_monthly_tax_on_worksheet(worksheet, current_row, item)
    return

def adjust_chargeable_monthly_tax_on_worksheet(worksheet, row, item):
    relief_column = 'N'
    chargeable_tax_column = 'M'
    chargeable_tax = item + 2400
    salary_value = worksheet['{column}{row}'.format(row=row, column='C')].value

    if chargeable_tax <= 2400:
        if salary_value == 0: # not working in that month
            worksheet['{column}{row}'.format(column=chargeable_tax_column, row=row)] = 0
            worksheet['{column}{row}'.format(column=relief_column, row=row)] = 0
        else: # low salary, not taxable
            worksheet['{column}{row}'.format(column=chargeable_tax_column, row=row)] = 2400
    else:
        worksheet['{column}{row}'.format(column=chargeable_tax_column, row=row)] = chargeable_tax
    return

def convert_files_to_pdf():
    os.system('libreoffice --headless --convert-to pdf ./repo/*.xlsx --outdir ./repo_pdf')
    return

if __name__ == '__main__':
    workbook = open_workbook()
    worksheet = get_records_worksheet(workbook)
    # max_row = 460
    iterate_through_rows(
        worksheet=worksheet, 
        min_row=5, 
        max_row=460,
        min_col=1, 
        max_col=28
    )
    convert_files_to_pdf()