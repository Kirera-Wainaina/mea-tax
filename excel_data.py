#  filename: excel1.py

import openpyxl
import shutil
from openpyxl.drawing.image import Image
from openpyxl.styles.borders import Border, Side

"""
I am trying to copy data from
bookOfExpenses.xlsx to
a copy of a p9 template

"""
#  function to open a spreadsheet copy contents
#  and input them in someone's file


class ExcelFigures():

	def __init__(self, min_col=1, max_col=33, min_row=1, max_row=466):
		self.min_col = min_col
		self.max_col = max_col
		self.min_row = min_row
		self.max_row = max_row


class ExcelFiles():
	def __init__(self, file):
		self.newFileNames = []
		self.file = openpyxl.load_workbook(file)
		self.fileSheet = self.file[self.file.sheetnames[0]]

	def createEverybodysFile(self):
		#  file is the name with the names
		#  file = 'mainsheet.xlsx'
		abspath_repo = '/home/richard/Documents/DadTx/repo/'
		for number in range(ExcelFigures().min_row + 4, ExcelFigures().max_row + 1):
			self.newFileNames.append(self.fileSheet.cell(row=number, column=2).value)

		for name in self.newFileNames:
			destination = abspath_repo + '%s.xlsx' % name
			shutil.copy('/home/richard/Documents/DadTx/static/p9.xlsx', destination)
		
	def transferIdentifyingInfo(self):
		# open both files and transfer from one to the other
		min_row = ExcelFigures().min_row + 4
		while min_row <= ExcelFigures().max_row:
			# 'while' allows going one row at a time
			for pin, name in self.fileSheet.iter_rows(min_col=ExcelFigures().min_col,
							  max_col=ExcelFigures().min_col + 1,
							  min_row=min_row,
							  max_row=min_row,
							  values_only=True):
				book = openpyxl.load_workbook('/home/richard/Documents/DadTx/repo/%s.xlsx' % name)
				self.bookSheet = book.active
				self.bookSheet['D12'] = name
				self.bookSheet['L14'] = pin
				img = Image('/home/richard/Documents/DadTx/static/p9_logo.png')
				self.bookSheet.add_image(img, 'h2')

				book.save(filename='/home/richard/Documents/DadTx/repo/%s.xlsx' % name)
				min_row += 1

	def transferFinancialInfo(self):

		min_row = ExcelFigures().min_row + 4
		while min_row <= ExcelFigures().max_row:
		# 'While' allows going one row at a time
			salaryList = []
			for salary in self.fileSheet.iter_cols(min_col=ExcelFigures().min_col + 2,
								max_col=ExcelFigures().max_col - 7,
								min_row=min_row,
								max_row=min_row,
								values_only=True):
				salary = list(salary)
				salaryList += salary
				#  the salaryList now has the list of salary and taxes


			book = openpyxl.load_workbook('/home/richard/Documents/DadTx/repo/%s.xlsx' %
							self.newFileNames[min_row - 5])
			# 'min_row-5' allows the first name in excel to be index 0
			bookSheet = book.active
			salary_row = 26
			tax_row = 26
			# the two above variables will help with distributing the figures
			for index in range(len(salaryList)):
			# the above variable will help with moving the salaries
			# row after row
				if index % 2 == 0:
				#  if divisible by 2 then that is salary
				#  transfer the salary to the persons template
					if salaryList[index] == "-" or salaryList[index] == 0:
						# some salaries are dashes in the records file
						# resulting in a #value! error
						bookSheet['C%s' % salary_row] = 0
						# if someone didn't earn a salary then tax(M) and relief(N)
						# should be 0 for those months
						bookSheet['M%s' % salary_row] = 0
						bookSheet['N%s' % salary_row] = 0
					else:
						bookSheet['C%s' % salary_row] = salaryList[index]
					salary_row += 1
				
				else:
					#  the index is odd
					#  this one goes to the PAYE column(O)
					if salaryList[index] == "-":
						bookSheet['O%s' % tax_row] = 0
					else:
						bookSheet['O%s' % tax_row] = salaryList[index]

					# tax charge column (M)
					if salaryList[index] == '-' or salaryList[index] == None:
					# at this point tax is 0, confirm if salary is 0 as well
						if bookSheet['C%s' % tax_row].value == 0:
							taxCharge = 0
						else:
							taxCharge = 0 + 2400
						bookSheet['M%s' % tax_row] = taxCharge
					else:
						taxCharge = float(salaryList[index]) + 2400
						bookSheet['M%s' % tax_row] = taxCharge
					tax_row += 1

			# thin = Side(border_style="thin",color='FF0000')
			# D28 = bookSheet.cell(row=29, column=4)
			# D28.value = "Richke"
			# D28.border = Border(top=thin, left=thin, right=thin, bottom=thin)
			# D28.border = Border(top=Side(border_style="thin", color="FF0000"))
	
			book.save(filename='/home/richard/Documents/DadTx/repo/%s.xlsx' % self.newFileNames[min_row - 5])
			min_row += 1

		
		
def main():
	start = ExcelFiles('/home/richard/Documents/DadTx/static/records.xlsx')
	start.createEverybodysFile()
	start.transferIdentifyingInfo()
	start.transferFinancialInfo()

	
main()
