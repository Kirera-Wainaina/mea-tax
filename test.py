
'''
PSEUDOCODE
    * Create a copy of the p9 template, give it a person's name
    * Open the file
    * Save the data in the file
'''

import unittest
from unittest.mock import MagicMock, patch
import openpyxl
import main # main.py
import os

class TestGetPayeWorksheet(unittest.TestCase):

    def setUp(self):
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.create_sheet('PAYE2022')
        self.sheet = main.get_records_worksheet(self.workbook)

    def test_worksheet_isInstance(self):
        self.assertIsInstance(self.sheet, 
            openpyxl.worksheet.worksheet.Worksheet)

    def test_worksheet_name(self):
        self.assertEqual(self.sheet.title, 'PAYE2022')

class WorkSheet:

    def iter_rows():
        pass

class TestIterateThroughRows(unittest.TestCase):

    def setUp(self):
        self.worksheet = WorkSheet()
        self.worksheet.iter_rows = MagicMock()

    def test_isCalled_withargs(self):
        min_row, max_row, min_col, max_col =5, 468, 1, 28
        main.iterate_through_rows(self.worksheet, min_row=min_row, 
            max_row=max_row, min_col=min_col, max_col=max_col)
        self.worksheet.iter_rows.assert_called_with(min_row=min_row, 
            max_row=max_row, min_col=min_col, max_col=max_col)

class TestHandleEmployeeDetails(unittest.TestCase):
    pass    

class TestCreateEmployeeTaxFile(unittest.TestCase):

    def test_path_isReturned(self):
        path = main.create_employee_tax_file('richard')
        self.assertEqual(path, '{cwd}/repo/{name}.xlsx'.format(
            cwd=os.getcwd(), name='richard'))

class TestAddEmployeeNameToTheirWorksheet(unittest.TestCase):

    def test_name_isAdded(self):
        mock_worksheet = MagicMock()
        main.add_employee_name_to_their_worksheet(mock_worksheet, 'john')
        mock_worksheet.__setitem__.assert_called_with('D12', 'john');

class TestAddEmployeePinToTheirWorksheet(unittest.TestCase):

    def test_pin_isAdded(self):
        mock_worksheet = MagicMock()
        main.add_employee_pin_to_their_worksheet(mock_worksheet, 'A0001')
        mock_worksheet.__setitem__.assert_called_with('L14', 'A0001')

class TestAddKraLogoToEmployeeWorksheet(unittest.TestCase):

    def test_image_isAdded(self):
        mock_worksheet = MagicMock()
        main.add_kra_logo_to_employee_worksheet(mock_worksheet)
        mock_worksheet.add_image.assert_called()

class TestAddFinancialDetailsToEmployeeWorksheet(unittest.TestCase):

    def setUp(self):
        self.mock_details = MagicMock()
        self.mock_worksheet = MagicMock()

    def test_financialDetailsVariable_isCreated(self):
        mock_details = MagicMock()
        main.add_financial_details_to_employee_worksheet(self.mock_worksheet, 
            self.mock_details)
        self.mock_details.__getitem__.assert_called()

class TestAddSalaryToEmployeeWorksheet(unittest.TestCase):

    def setUp(self):
        self.mock_worksheet = MagicMock()

    def test_worksheet_setsValue(self):
        main.add_salary_to_employee_worksheet(self.mock_worksheet, 2, 1500)
        self.mock_worksheet.__setitem__.assert_called_with('C27', 1500)

    def test_worksheet_getsZeroIfDash(self):
        main.add_salary_to_employee_worksheet(self.mock_worksheet, 4, '-')
        self.mock_worksheet.__setitem__.assert_called_with('C28', 0)

    def test_worksheet_addsZeroIfZero(self):
        main.add_salary_to_employee_worksheet(self.mock_worksheet, 6, 0)
        self.mock_worksheet.__setitem__.assert_called_with('C29', 0)

class TestAddTaxToEmployeeWorksheet(unittest.TestCase):
    
    def setUp(self):
        self.mock_worksheet = MagicMock()

    def test_worksheet_setsValue(self):
        main.add_tax_to_employee_worksheet(self.mock_worksheet, 1, 2600)
        self.mock_worksheet.__setitem__.assert_called_with('M26', 2600)

    def test_worksheet_getsZeroIfDash(self):
        main.add_tax_to_employee_worksheet(self.mock_worksheet, 3, '-')
        self.mock_worksheet.__setitem__.assert_any_call('M27', 0)

class TestAdjustFinalMonthlyTaxOnWorksheet(unittest.TestCase):

    def setUp(self):
        self.mock_worksheet = MagicMock()

    def test_monthlyTax_lessThanZero(self):
        main.adjust_final_monthly_tax_on_worksheet(self.mock_worksheet, 26, 1300)
        self.mock_worksheet.__setitem__.assert_called_with('O26', 0)

if __name__ == '__main__':
    unittest.main()